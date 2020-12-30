import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill,Border,colors,Side,Alignment
import os
import shutil
import xlsxwriter

path = './'
template_path = './analysis_module/performance/'
topath = './analysis_module/performance/performance.xlsx'
checkpath = '././analysis_module/performance/checklist_performance.txt'
temporary_file = './analysis_module/performance/temp/'

def judge_file(path,topath):
    if not os.path.exists(template_path):
        os.mkdir(template_path,1)
    if not os.path.exists(topath):
        if 'template_performance.xlsx' in os.listdir(path):
            shutil.copy('template_performance.xlsx',template_path)
            os.rename(template_path + 'template_performance.xlsx',template_path + 'performance.xlsx')
    if not os.path.exists(checkpath):
        with open(checkpath,mode='w',encoding='utf-8') as f:
            pass
    if not os.path.exists(temporary_file):
        os.mkdir(temporary_file,1)

def file(path,topath):
    c = 1
    dir_name = os.listdir(path)
    namelist = []
    with open(checkpath,encoding='utf-8') as cf:
        lines = cf.readlines()
        for line in lines:
            namelist.append(line.strip('\n'))
    cf.close()
    for n in range(len(dir_name)):
        sub_path = os.path.join(path,dir_name[n])
        filename = str(sub_path).split('/')[1]
        if os.path.isdir(sub_path):
            sub_fname = os.listdir(sub_path)
            if 'analysis' in sub_fname:
                if 'analysis.xlsx' in os.listdir(sub_path+'/analysis/'):
                    if filename in namelist:
                        print(filename + ' have done')
                        continue
                    else:
                        with open(checkpath,'a+',encoding='utf-8') as cf1:
                            cf1.write(filename + '\n')
                        cf1.close()

                    data = pd.read_excel(sub_path+'/analysis/analysis.xlsx',sheet_name="index")
                    df_data = pd.DataFrame(data)                        
                    title = '2020-' + str(sub_path).split("/")[-1] + '_集成测试_' + df_data['android_ver'][0]
                    wb = load_workbook(topath)
                    ws = wb["app_performance"]
                    #判断单元格是否为空
                    #从第11列开始遍历step = 6,找到最近一个为空的单元格
                    num = 1
                    for t in range(11,30000,6):
                        if ws.cell(1,t).value is None:
                            num = t
                            break
                    
                    ws.cell(1,num).value = title
                    ws.cell(2,num).value= 'RAM(M)'
                    ws.cell(2,num+1).value= 'CPU(%)'
                    ws.cell(2,num+2).value= 'APP Size(M)'
                    ws.cell(2,num+3).value= 'StartupTime(s)'
                    ws.cell(2,num+4).value= 'IO'
                    ws.cell(2,num+5).value= '显存(M)'

                    for j in range(2,len(ws['B'])):
                        for k in range(len(df_data['packages'])):
                            if df_data['packages'][k] == ws['B'][j].value:
                                ws[j+1][num-1].value = df_data['totalmem max'][k]
                                ws[j+1][num].value = df_data['cpu max'][k]
                    wb.save(topath)   

                    data2 = pd.read_excel(sub_path+'/analysis/analysis.xlsx',sheet_name="summary_capture")
                    df_data2 = pd.DataFrame(data2)
                    wb_sheet2 = load_workbook(topath)
                    ws2 = wb_sheet2["system_performance"]
                    r = 1
                    for r0 in range(1,60000):
                        if ws2[r0][2].value is None:
                            r = r0
                            break
                    ws2[r][c].value = title

                    for m in range(len(df_data2["CAPTURE_TYPE"])):
                        if df_data2["CAPTURE_TYPE"][m] == "capture_cpu_idle_low":
                            ws2[r][c+1].value = 300 - int(df_data2["%idle"][m])
                        if df_data2["CAPTURE_TYPE"][m] == "capture_mem_available_low":
                            ws2[r][c+2].value = df_data2["Available_RAM"][m]
                    wb.save(topath)

                    for m in range(len(df_data2["CAPTURE_TYPE"])):
                        if df_data2["CAPTURE_TYPE"][m] == "capture_cpu_idle_low":
                            ws2[r][c+1].value = 300 - int(df_data2["%idle"][m])
                        if df_data2["CAPTURE_TYPE"][m] == "capture_mem_available_low":
                            ws2[r][c+2].value = df_data2["Available_RAM"][m]
                    wb_sheet2.save(topath) 
                else:
                    print(str(sub_path).split("/")[-1]+" not exist ananlysis.xlsx 文件")
                    continue
                if 'analysis_meminfo.csv' in os.listdir(sub_path+'/analysis/'):
                    mem_data = pd.read_csv(sub_path+'/analysis/analysis_meminfo.csv',encoding='utf-8')
                    df_mem = pd.DataFrame(mem_data)

                    wb_sheetname2 = load_workbook(topath)
                    ws_sheet2 = wb_sheetname2["system_performance"]
                    ws_sheet2[r][c+3].value = float(df_mem["non_contig_len(MB)"].min())
                    r += 1
                    wb_sheetname2.save(topath) 
                    print(str(sub_path).split("/")[-1] + " performance data import completed")
                else:
                    print(str(sub_path).split("/")[-1] + " lack non_contig_len(MB) data")
                    print(str(sub_path).split("/")[-1] + " performance data import completed")
                    continue              
            else:
                continue


#给指定区域设置粗匣框线/设置单元格背景颜色
def set_solid_border(topath):

    wb = load_workbook(topath)
    sheet = wb["app_performance"]
    rows = sheet.max_row
    cols = sheet.max_column
    fill0 = PatternFill('solid',fgColor='9BC2E6')
    fill1 = PatternFill('solid',fgColor='C6E082')
    fill2 = PatternFill('solid',fgColor='FFE699')

    for col in range(1,cols+1):
        sheet.cell(row=1,column=col).fill = fill0
        sheet.cell(row=2,column=col).fill = fill1

    line_m = Side(style='medium', color='000000')  # 粗边框
    border1 = Border(top=line_m)
    border2 = Border(right=line_m)
    border3 = Border(top=line_m,right=line_m)
    border4 = Border(bottom=line_m)
    border5 = Border(bottom=line_m,right=line_m)

    for r in range(1,rows+1):
        for c in range(1,cols+1):
            if r == 3:
                sheet.cell(r,c).border = border1
            if c == 1 or (c-4)%6 == 0:
                sheet.cell(r,c).border = border2
            if r == 3 and (c-4)%6 == 0 or r == 3 and c == 1:
                sheet.cell(r,c).border = border3
            if r == rows:
                sheet.cell(r,c).border = border4
            if r == rows and (c-4)%6 == 0 or r == rows and (c-4)%6 == 0 or\
               r == rows and c == 1 or r == rows and c == 1:
               sheet.cell(r,c).border = border5         
    wb.save(topath)

    wb_sheet2 = load_workbook(topath)
    ws2 = wb_sheet2["system_performance"]
    align = Alignment(horizontal='center')
    ws2.merge_cells('C1:E1')
    ws2['C1'].alignment = align

    for row in range(1,ws2.max_row+1):
        ws2.cell(column=1,row=row).fill = fill2
    for col in range(1,ws2.max_column+1):
        ws2.cell(row=1,column=col).fill = fill0
        ws2.cell(row=2,column=col).fill = fill1
    
    for line in range(1,ws2.max_row+1):
        for ver_line in range(1,ws2.max_column+1):
            if line == 3:
                ws2.cell(line,ver_line).border = border1
            if ver_line == 1 or ver_line == ws2.max_column:
                ws2.cell(line,ver_line).border = border2
            if line == 3 and ver_line == 1 or line == 3 and ver_line == ws2.max_column:
                ws2.cell(line,ver_line).border = border3
            if line == ws2.max_row:
                ws2.cell(line,ver_line).border = border4
            if line == ws2.max_row and ver_line == 1 or line == ws2.max_row and ver_line == ws2.max_column:
                ws2.cell(line,ver_line).border = border5
    wb_sheet2.save(topath)


#筛选数据
def filter_RAM(topath):
    wb = load_workbook(topath)
    new_ws = wb['exceeding_app_ram']
    #找到第一行不为空的最大列对应的值
    for r in range(1,new_ws.max_column+1):
        if new_ws.cell(1,r).value is not None:
            reference = new_ws.cell(1,r).value
    # print(reference)
    ws = wb['app_performance']
    cols = []
    re = 1
    if reference == 'RAM(M)':
        for c0 in range(1,ws.max_column):
            if ws.cell(2,c0).value == 'RAM(M)':
                cols.append(c0-1)
    else:
        for rr in range(1,ws.max_column):
            if ws.cell(1,rr).value == reference:
                re = rr + 6
            else:
                continue
    
        cols = [4]
        for c in range(re,ws.max_column):
            if ws.cell(2,c).value == 'RAM(M)':
                cols.append(c-1)
                
    wb.save(topath)
    data = pd.read_excel(topath,usecols=cols)
    df_data = pd.DataFrame(data)
    standard = df_data.iloc[1:,0:1]
    need_df = df_data.iloc[1:,1:,]/1024
    version = []
    if len(cols) > 1:
        for nc in range(cols[1]+1,ws.max_column,6):
            if ws.cell(1,nc).value is not None:
                version.append(ws.cell(1,nc).value)
    else:
        pass

    for i in range(len(standard)):
        for j in range(len(need_df.columns)):
            if need_df.iloc[i,j] < standard.iloc[i,0]:
                need_df.iloc[i,j] = ""

    new_df = list(need_df.columns)
    mc = 0
    for lc in range(len(new_df)):
        new_df[lc] = version[mc]
        if mc < len(version):
            mc += 1
    need_df.columns = new_df
    need_df.to_excel(temporary_file + 'test.xlsx')
    cp_file = load_workbook(temporary_file + 'test.xlsx')
    cp_ws = cp_file['Sheet1']
    #找第一行第一个为空的单元格
    num = 1
    for t in range(1,30000):
        if new_ws.cell(1,t).value is None:
            num = t
            break

    for n in range(1,cp_ws.max_column):
        for m in range(1,cp_ws.max_row+1):
            new_ws.cell(m,num).value = cp_ws[m][n].value
        num +=1 
    wb.save(topath)
    # print("RAM data filtering completed")
    os.remove(temporary_file + 'test.xlsx')


def filter_CPU(topath):
    wb = load_workbook(topath)
    new_ws = wb['exceeding_app_cpu']
    #找到第一行不为空的最大列对应的值
    for r in range(1,new_ws.max_column+1):
        if new_ws.cell(1,r).value is not None:
            reference = new_ws.cell(1,r).value

    ws = wb['app_performance']
    cols = []
    re = 1
    if reference == 'CPU(%)':
        for cc0 in range(1,ws.max_column):
            if ws.cell(2,cc0).value == 'CPU(%)':
                cols.append(cc0-1)
    else:
        for rr in range(1,ws.max_column):
            if ws.cell(1,rr).value == reference:
                re = rr + 6
            else:
                continue
        cols = [5]
        for cc in range(re,ws.max_column):
            if ws.cell(2,cc).value == 'CPU(%)':
                cols.append(cc-1)
    wb.save(topath)

    data = pd.read_excel(topath,usecols=cols)
    df_data = pd.DataFrame(data)
    standard = df_data.iloc[1:,0:1]*100
    need_df = df_data.iloc[1:,1:,]
    version = []
    if len(cols) > 1:
        for nc in range(cols[1],ws.max_column,6):
            if ws.cell(1,nc).value is not None:
                version.append(ws.cell(1,nc).value)
    else:
        pass
    
    for i in range(len(standard)):
        for j in range(len(need_df.columns)):
            if need_df.iloc[i,j] < standard.iloc[i,0]:
                need_df.iloc[i,j] = ""
    
    new_df = list(need_df.columns)
    mc = 0
    for lc in range(len(new_df)):
        new_df[lc] = version[mc]
        if mc < len(version):
            mc += 1
    need_df.columns = new_df
    need_df.to_excel(temporary_file + 'test.xlsx')
    cp_file = load_workbook(temporary_file + 'test.xlsx')
    cp_ws = cp_file['Sheet1']
    #找第一行第一个为空的单元格
    num = 1
    for t in range(1,30000):
        if new_ws.cell(1,t).value is None:
            num = t
            break

    for n in range(1,cp_ws.max_column):
        for m in range(1,cp_ws.max_row+1):
            new_ws.cell(m,num).value = cp_ws[m][n].value
        print()
        num +=1 
    wb.save(topath)
    # print('CPU data filtering completed')
    os.remove(temporary_file + 'test.xlsx')
    os.rmdir(temporary_file)
    

if __name__ == "__main__":
    judge_file(path,topath)
    file(path,topath)
    set_solid_border(topath)
    filter_RAM(topath)
    filter_CPU(topath)
