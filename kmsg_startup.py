# version 0.0.2
import os
import shutil
import pandas as pd
import sys
import re
import linecache
import csv
import zipfile
from openpyxl import load_workbook
import xlrd,xlwt
from openpyxl import Workbook
import operator
import datetime,time
import shutil


#处理所有日期目录下的kmsg log
class Clear_KMSG():
    fpath = './'
    temporary_file = './analysis_module/performance/temp/'
    __cleared_file = './analysis_module/performance/startupTimeFromKmsg/'
    analysis_path = './analysis_module'
    total_path = './analysis_module/performance/'
    file_path = total_path + '/' + 'startupTimeFromKmsg.xlsx'
    key = 'Kmsg'
    #已处理过的将文件名写入date_checklist.txt文件中
    date_check_path = './analysis_module/performance/checklist_startup.txt'
 
    if not os.path.exists(analysis_path):
        os.mkdir(analysis_path)
    if not os.path.exists(total_path):
        os.mkdir(total_path)
    if not os.path.exists(__cleared_file):
        os.mkdir(__cleared_file)
    if os.path.exists(temporary_file):
        ls = os.listdir(temporary_file)
        for l in ls:
            f_path = os.path.join(temporary_file,l)
            if os.path.isfile(f_path):
                os.remove(f_path)
    else:
        os.mkdir(temporary_file)
    if not os.path.exists(date_check_path):
        with open(date_check_path,mode='w',encoding='utf-8') as f:
            pass   
    reference = []
    with open(date_check_path,encoding='utf-8') as cf:
        lines = cf.readlines()
        for line in lines:
            reference.append(line.strip('\n'))
    cf.close()
        

    def judge_file(self):
        file = os.listdir(self.fpath)
        file.sort()
        new_date= []
        for t in range(len(file)):
            sub_file = os.path.join(self.fpath,file[t])
            if os.path.isdir(sub_file):
                filename = os.listdir(sub_file)
                filename.sort()
                if 'log' in filename:
                    sub_path = os.path.join(sub_file + '/log')
                    sub_filename = os.listdir(sub_path)
                    sub_filename.sort()
                    # print(sub_filename)
                    if 'kmsg' in sub_filename:
                        date = str(sub_file).split('/')[1]
                        if date in self.reference:
                            print(date + ' has already created, pass')
                            continue
                        else:
                            with open(self.date_check_path,'a+',encoding='utf-8') as cf1:
                                new_date.append(date)
                                cf1.write(date + '\n')   
                            cf1.close()
                            sub_filepath = os.path.join(sub_path + '/kmsg')
                            used_file = os.listdir(sub_filepath)
                            used_file.sort()
                            for name in used_file:
                                if self.key in name and name.split(".")[1] == "txt":
                                    zip_path = sub_filepath + '/' + name
                                    shutil.copy(zip_path, self.temporary_file + '/' + name)
                            used_path = []
                            for xname in os.listdir(self.temporary_file):
                                final_path = os.path.join(self.temporary_file, xname)
                                if os.path.getsize(final_path) == 0:
                                    os.remove(final_path)
                                else:
                                    used_path.append(final_path)
                            
                            if os.path.exists(sub_file + '/android_ver.txt'):
                                with open(sub_file + '/android_ver.txt') as f:
                                    version = f.read()
                                f.close()
                            else:
                                version = ""

                            key_list = []
                            with open('./config_startup_keyword.txt') as kf:
                                key_words = kf.readlines()
                                for word in key_words:
                                    key_list.append(word.strip("\n"))
                            # print(key_list)
                            path = []
                            if used_path != []:
                                for j in range(len(used_path)): 
                                    with open(used_path[j],encoding='utf-8') as f:
                                        content = f.read()
                                        if key_list[0] in content:
                                            continue
                                        if key_list[1] in content:
                                            path.append(used_path[j])
                                        else:
                                            continue
                                # print(path)
                            else:
                                print('no file')

                            v_dict = {}
                            if path != []:    
                                v_dict['date'] = date
                                v_dict['version'] = version
                                v_dict['filename'] = str(path[0]).split('/')[2]
                            else:
                                v_dict['filename'] = ""

                            dict = {}
                            for m in range(len(key_list)):
                                dict[key_list[m]] = []
                            if len(path) == 0:
                                print(date + " no information")
                                for p in used_path:
                                    os.remove(p)
                                continue
                            elif len(path) == 1:
                                # for k in range(len(path)):
                                with open(path[0],encoding='utf-8') as endfile1:
                                    lines = endfile1.readlines()
                                    for line in lines:
                                        for i in range(len(key_list)):
                                            if key_list[i] in line:
                                                dict[key_list[i]].append(int(line.split(",")[2])/1000000)
                                    for n in list(dict.keys()):
                                        if len(dict[n]) > 1:
                                            dict[n]=[dict[n][0]]

                                with open(self.__cleared_file + date + '.csv','w',encoding='utf-8',newline="") as wf1:
                                    fieldnames = ['keyword','content']
                                    writer = csv.DictWriter(wf1,fieldnames=fieldnames)
                                    writer.writeheader()
                                    csv_data = csv.writer(wf1)
                                    for key,value in v_dict.items():
                                        csv_data.writerow([key]+[value])
                                    for key,value in dict.items():
                                        csv_data.writerow([key]+value)      
                                wf1.close()
                                for p in used_path:
                                    os.remove(p)
                            else:
                                filepath = self.__cleared_file + date + '.csv'
                                with open(filepath,'w',encoding='utf-8',newline="") as endfile2:
                                    fieldnames = ['keyword']
                                    for number in range(1,len(path)+1):
                                        fieldnames.append('content_{}'.format(number))
                                    writer = csv.DictWriter(endfile2,fieldnames=fieldnames)
                                    writer.writeheader()
                                endfile2.close()

                                with open(filepath,'a',newline="") as wf2:
                                    csv_write = csv.writer(wf2,dialect='excel')
                                    csv_write.writerow(['date'])
                                    csv_write.writerow(['version'])
                                    csv_write.writerow(['filename'])
                                    for k in range(len(key_list)):
                                        csv_write.writerow([key_list[k]])
                                wf2.close()
                                
                                new_dict = {}
                                for k in range(len(path)):
                                    new_dict['date'] = [date + '_{}'.format(k)]
                                    new_dict['version'] = [version]
                                    new_dict['filename'] = []

                                    for v in range(0,len(key_list)):
                                        new_dict[key_list[v]] = []

                                    with open(path[k],encoding='utf-8') as endfile3:
                                        new_dict['filename'].append(str(path[k].split('/')[2]))
                                        lines = endfile3.readlines()
                                        for line in lines:
                                            for num in range(0,len(key_list)):
                                                if key_list[num] in line:
                                                    new_dict[key_list[num]].append(int(line.split(",")[2])/1000000)
                                        for key in list(new_dict.keys()):
                                            if len(new_dict[key]) > 1:
                                                new_dict[key] = [new_dict[key][0]]

                                    data = pd.read_csv(filepath,encoding='utf-8')
                                    df_data = pd.DataFrame(data)
                                    for c in new_dict.keys():
                                        for v in range(len(df_data['keyword'])):
                                            if df_data['keyword'][v] == c :
                                                if new_dict[c] != []:
                                                    # df_data[fieldnames[k+1]][v] = new_dict[c][0]
                                                    df_data.loc[v,fieldnames[k+1]] = new_dict[c][0]
                                                elif new_dict[c] == []:
                                                    # df_data[fieldnames[k+1]][v] = ""
                                                    df_data.loc[v,fieldnames[k+1]] = ""
                                            else:
                                                continue
                                    df_data.to_csv(filepath,index=False)
                                for p in used_path:
                                    os.remove(p)
                        print(date + ' has done')
        return new_date

    
    def create_csv(self,num=0):
        wb = Workbook()
        ws = wb.active
        ws1 =wb.create_sheet('总表',0)
        key_list = []

        with open('config_startup_keyword.txt') as kf:
            key_words = kf.readlines()
            for word in key_words:
                key_list.append(word.strip("\n"))
        # print(key_list)
        ws_used = wb['总表']
        ws_used.cell(1,1).value = 'date'
        ws_used.cell(2,1).value = 'version'
        ws_used.cell(3,1).value = 'filename'
        for i in range(4,len(key_list)+4):
            ws_used.cell(i,1).value = key_list[i-4]

        filename = os.listdir(self.__cleared_file)
        filename.sort()
        for j in range(len(filename)):
            file_data = pd.read_csv(self.__cleared_file + filename[j],encoding='gbk')
            df_data = pd.DataFrame(file_data)
            if len(df_data.columns) == 2:
                for m in range(1,len(key_list)+4):
                    ws_used.cell(m,2+num).value = df_data['content'][m-1]
                num += 1
            if len(df_data.columns) > 2:
                for v in range(1,len(df_data.columns)):
                    for n in range(1,len(key_list)+4):
                        ws_used.cell(n,2+num).value = df_data[df_data.columns[v]][n-1]
                    num += 1
        # print(num)
        wb.save(self.file_path)
        wb.close()

    
    def add_csv(self,num=1):
        file = os.listdir(self.fpath)
        file.sort()
        new_date= []
        for t in range(len(file)):
            sub_file = os.path.join(self.fpath,file[t])
            if os.path.isdir(sub_file):
                filename = os.listdir(sub_file)
                filename.sort()
                if 'log' in filename:
                    sub_path = os.path.join(sub_file + '/log')
                    sub_filename = os.listdir(sub_path)
                    sub_filename.sort()
                    # print(sub_filename)
                    if 'kmsg' in sub_filename:
                        date = str(sub_file).split('/')[1]
                        if date in self.reference:
                            continue
                        else:
                            with open(self.date_check_path,'a+',encoding='utf-8') as cf1:
                                new_date.append(date)
                            cf1.close()

        wb = load_workbook(self.file_path)
        ws = wb['总表']
        #从第一列开始找到最近的一个为空的单元格
        num = 1
        for t in range(1,30000):
            if ws.cell(1,t).value is None:
                num = t
                break

        #从这一列开始续写新日期所包含的数据内容
        filename = os.listdir(self.__cleared_file)
        filename.sort()
        deal_fname = []
        for name in filename:
            if name.split(".")[0] in new_date:
                deal_fname.append(name)
        if len(deal_fname) != []:
            for j in range(len(deal_fname)):
                file_data = pd.read_csv(self.__cleared_file + deal_fname[j],encoding='gbk')
                df_data = pd.DataFrame(file_data)
                if len(df_data.columns) == 2:
                    for m in range(1,ws.max_row+1):
                        ws.cell(m,num).value = df_data['content'][m-1]
                    num += 1
                if len(df_data.columns) > 2:
                    for v in range(1,len(df_data.columns)):
                        for n in range(1,ws.max_row+1):
                            ws.cell(n,num).value = df_data[df_data.columns[v]][n-1]
                        num += 1
        wb.save(self.file_path)
        wb.close()
        

    def deal_csv(self):
        if 'startupTimeFromKmsg.xlsx' in  os.listdir(self.total_path):
            file = pd.read_excel(self.file_path,sheet_name='总表')
            file_df = pd.DataFrame(file)
            last_key = []

            for i in range(1,len(file_df['date'])):
                last_key.append(file_df['date'][i])
            # print(last_key)
            newkey_list = ['date','version','filename']
            with open('./config_startup_keyword.txt') as new_kf:
                new_keys = new_kf.readlines()
                for word in new_keys:
                    newkey_list.append(word.strip("\n"))
            # print(newkey_list)
            if operator.eq(last_key,newkey_list[2:]) is True:
                clear_kmsg.add_csv()
            else:
                #若关键字改变则重命名sheetname,并创建新的sheet ['总表']
                last_wb = load_workbook(self.file_path) 
                last_ws = last_wb['总表'] 
                time = str(datetime.datetime.today()).split(" ")[0]
                last_ws.title = "总表_{}".format(time)
                last_wb.create_sheet(title='总表',index=0)
                current_ws_sheet1 = last_wb['总表']
                current_ws_sheet2 = last_wb[last_ws.title]
                #旧关键字表中与新关键字表相同，则将所在行写到新的sheet['总表']中，否则将新增关键字写到新‘总表’A列
                row0 = 1
                for j in range(len(newkey_list)):
                    for k in range(1,current_ws_sheet2.max_row+1):
                        if current_ws_sheet2.cell(k,1).value  == newkey_list[j]:
                            for l in range(1,current_ws_sheet2.max_column+1):
                                current_ws_sheet1.cell(row0,l).value = current_ws_sheet2.cell(k,l).value
                            break
                        else:
                            continue
                    current_ws_sheet1.cell(row0,1).value = newkey_list[j]
                    row0 += 1
                last_wb.save(self.file_path)
                last_wb.close()
                #旧数据处理完毕后，在新的sheet‘总表’中开始写入新关键字提取的数据内容
                clear_kmsg.add_csv()
        else:
            clear_kmsg.create_csv()
        os.removedirs(self.temporary_file)


def start_report_startup():
    global clear_kmsg
    clear_kmsg = Clear_KMSG()
    clear_kmsg.judge_file()
    clear_kmsg.deal_csv()
    # os.removedirs('./file/')
    # del clear_kmsg


if __name__ == '__main__':
    start_report_startup()




